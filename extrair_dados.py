r"""
Extrai os dados do painel "Melhoria Salarial" a partir de RESULTADO.xlsm
(pasta MELHORIA SALARIO) e salva um dados.json pronto pro gerador de HTML
consumir.

A planilha original (abas ITENS FOCO / RESUMO) só calcula 1 RCA por vez —
o código digitado em ITENS FOCO!C8 decide quem aparece. O painel gerado
reproduz esse MESMO jeito de trabalhar (digita o código, vê só aquele
vendedor), então este script exporta os dados BRUTOS de TODOS os RCAs
(pra não precisar reabrir o Excel a cada troca de vendedor) e deixa toda
a matemática (comissão, metas, checkboxes) pro JavaScript do painel:

- '315' (ROTINA315): 1 linha por RCA — nome (col B), total de pedidos
  (col G), valor vendido (col J).
- BACON / CALABRESA / FRESCAIS / PAES / LACTIOS / BATATA / BOVINO / SUINO:
  1 linha por RCA — peso (col K), valor realizado (col L), positivação
  (col D).

Taxa de comissão por categoria é fixa pra empresa toda (confirmado lendo
a fórmula de cada bloco, 21/08): bacon=3% calabresa=3% frescais=3%
paes=3% lactios=3% batata=3% bovino=1,5% suino=3% — exportada em
constantes.taxas_categoria.

Meta de positivação por categoria (I na ITENS FOCO), meta de pedidos/dia
(P17) e a taxa média de comissão sobre pedidos (O11) eram fixas na
planilha, mas viram CONFIGURAÇÃO EDITÁVEL no painel (o Edmar ajusta e
recalcula na hora pra qualquer vendedor) — aqui só exportamos os valores
atuais como ponto de partida, em constantes.metas_categoria_padrao /
meta_pedidos_dia / taxa_padrao.

Industrializado/Thermo/Positivação Dia 15/Dia 30/Prêmio Campanha usam
CHECKBOX MANUAL na planilha original (M17/M20/S11/S14/V11) — não tem
fórmula, é o Edmar revisando RCA por RCA. Aqui a gente só extrai o valor
"potencial" (o teto de cada bônus, se bater a meta); o "bateu ou não" fica
editável no próprio painel (checkbox no navegador, guardado no
localStorage por RCA — ver gerar_painel.py).
"""

import json
import os

import openpyxl

CAMINHO_RESULTADO = r"C:\Users\edmar\Desktop\MELHORIA SALARIO\RESULTADO.xlsm"
CAMINHO_PAINEL_PILARES = r"c:\AutomacaoMaxGestao\painel_pilares\dados.json"
CAMINHO_PAINEL_DEPARTAMENTOS = r"c:\AutomacaoMaxGestao\painel_departamentos\dados.json"

# Categoria daqui -> categoria equivalente no Painel Departamentos (mesmo
# produto, nome de chave diferente em cada planilha). "vegetais" não tem
# equivalente lá ainda (categoria nova, sem bloco de meta em Departamentos).
CATEGORIA_PARA_DEPARTAMENTOS = {
    "bacon": "bacon",
    "calabresa": "calabresa",
    "frescais": "frescais",
    "paes": "paes",
    "lactios": "lacteos",
    "batata": "batata",
    "bovino": "bovino",
    "suino": "suino",
    "thermo_cat": "thermo",
}

PASTA_BASE = os.path.dirname(os.path.abspath(__file__))
CAMINHO_SAIDA = os.path.join(PASTA_BASE, "dados.json")

# (chave, rótulo, aba de origem, taxa de comissão, meta de positivação)
CATEGORIAS = [
    ("bacon", "Bacon", "BACON", 0.03, 5),
    ("calabresa", "Calabresas", "CALABRESA", 0.03, 22),
    ("frescais", "Linguiças/Frescais", "FRESCAIS", 0.03, 15),
    ("paes", "Pães", "PAES", 0.03, 17),
    ("lactios", "Lácteos", "LACTIOS", 0.03, 23),
    ("batata", "Batata", "BATATA", 0.03, 15),
    ("bovino", "Bovino", "BOVINO", 0.015, 10),
    ("suino", "Suíno", "SUINO", 0.03, 6),
    ("thermo_cat", "Thermo", "THERMO", 0.03, 6),
    ("vegetais", "Vegetais", "VEGETAIS", 0.03, 6),
]

PREMIO_FIXO = 150  # teto de Dia 15 / Dia 30 / Campanha (mesmo valor pra todo mundo, confirmado em ITENS FOCO)

RECOMPRA_LIMITE = 0.20  # ITENS FOCO!Y10 = SE(X10<20%, 150, 0)
RECOMPRA_PREMIO = 150


def _num(v):
    return v if isinstance(v, (int, float)) else 0


def _ler_categoria(wb, nome_aba):
    """{codigo: {"peso":..,"valor":..,"positivacao":..}} — cada aba de
    categoria tem 1 linha por RCA, sem cabeçalho: A=código, K=peso,
    L=valor realizado, D=positivação."""
    ws = wb[nome_aba]
    out = {}
    for row in ws.iter_rows(values_only=True):
        codigo = row[0]
        if codigo is None:
            continue
        out[int(codigo)] = {
            "peso": _num(row[10]),      # K
            "valor": _num(row[11]),     # L
            "positivacao": _num(row[3]),  # D
        }
    return out


def _ler_recompra(wb):
    """Recompra é automática (não tem checkbox de 'bateu' na planilha) —
    calculada a partir da aba '8110' (1 linha por cliente atendido por
    RCA). % = quantos clientes fizeram só 1 pedido (QTPED=1, coluna Q) /
    total de clientes do RCA. Ganha R$150 se essa % ficar abaixo de 20%
    (ITENS FOCO!X10/Y10) — quanto menor, melhor (menos cliente "de
    passagem", mais gente recomprando)."""
    if "8110" not in wb.sheetnames:
        return {}
    ws = wb["8110"]
    total_por_rca = {}
    um_pedido_por_rca = {}
    primeiro = True
    for row in ws.iter_rows(values_only=True):
        if primeiro:
            primeiro = False
            continue  # cabeçalho
        codigo = row[0]
        if codigo is None:
            continue
        codigo = int(codigo)
        qtped = row[16] if len(row) > 16 else None  # Q
        total_por_rca[codigo] = total_por_rca.get(codigo, 0) + 1
        if qtped == 1:
            um_pedido_por_rca[codigo] = um_pedido_por_rca.get(codigo, 0) + 1
    return {
        codigo: (um_pedido_por_rca.get(codigo, 0) / total) if total else 0
        for codigo, total in total_por_rca.items()
    }


def _ler_supervisores():
    """Cross-referencia com o painel 4 Pilares pra saber o supervisor de
    cada RCA (essa planilha de melhoria salarial não tem essa coluna)."""
    if not os.path.exists(CAMINHO_PAINEL_PILARES):
        return {}
    with open(CAMINHO_PAINEL_PILARES, "r", encoding="utf-8") as f:
        dados = json.load(f)
    return {int(r["codigo"]): r["supervisor"] for r in dados}


def _ler_industrializado_pilares():
    """Cross-referencia com o painel 4 Pilares pra pegar a participação % e
    a margem % REAL de industrializado de cada RCA (a planilha de melhoria
    salarial só tem a META de industrializado, não o realizado)."""
    if not os.path.exists(CAMINHO_PAINEL_PILARES):
        return {}
    with open(CAMINHO_PAINEL_PILARES, "r", encoding="utf-8") as f:
        dados = json.load(f)
    return {
        int(r["codigo"]): {
            "participacao_pct": r["industrializado"]["participacao_pct"],
            "margem_pct": r["industrializado"]["margem_pct"],
        }
        for r in dados
    }


def _ler_metas_departamento():
    """Cross-referencia com o Painel Departamentos pra pegar a META de
    positivação de cada categoria — a mesma que o supervisor já acompanha
    lá (bloco do supervisor define 1 meta por categoria pra todos os RCAs
    dele). Chave de retorno já traduzida pro nome de categoria usado aqui
    (ver CATEGORIA_PARA_DEPARTAMENTOS)."""
    if not os.path.exists(CAMINHO_PAINEL_DEPARTAMENTOS):
        return {}
    with open(CAMINHO_PAINEL_DEPARTAMENTOS, "r", encoding="utf-8") as f:
        dados = json.load(f)
    out = {}
    for r in dados:
        metas = {}
        for chave_aqui, chave_dep in CATEGORIA_PARA_DEPARTAMENTOS.items():
            info = r["categorias"].get(chave_dep)
            if info:
                metas[chave_aqui] = info["meta"]
        out[int(r["codigo"])] = metas
    return out


def extrair():
    wb = openpyxl.load_workbook(CAMINHO_RESULTADO, data_only=True)
    ws_geral = wb["ITENS FOCO"]

    dias_uteis = _num(ws_geral["O14"].value) or 23
    meta_pedidos_dia = _num(ws_geral["P17"].value) or 15
    taxa_padrao = _num(ws_geral["O11"].value) or 0.0178

    cache_categorias = {chave: _ler_categoria(wb, aba) for chave, _, aba, _, _ in CATEGORIAS}
    supervisores = _ler_supervisores()
    industrializado_pilares = _ler_industrializado_pilares()
    recompra_por_rca = _ler_recompra(wb)
    metas_departamento = _ler_metas_departamento()

    ws315 = wb["315"]
    rcas = []
    for row in ws315.iter_rows(values_only=True):
        codigo = row[0]
        if codigo is None:
            continue
        codigo = int(codigo)
        nome_bruto = str(row[1] or "").strip()
        if " - " in nome_bruto:
            nome_rca, rota = nome_bruto.split(" - ", 1)
        else:
            nome_rca, rota = nome_bruto, ""

        total_pedidos = _num(row[6])   # G
        valor_vendido = _num(row[9])   # J

        # Só os dados brutos por categoria (peso/valor/positivação) — a
        # comissão atual/potencial é calculada no navegador, porque a meta
        # de positivação e a meta de pedidos agora são editáveis ali.
        categorias_dados = {}
        for chave, _label, _aba, _taxa, _meta_posit in CATEGORIAS:
            info = cache_categorias[chave].get(codigo, {"peso": 0, "valor": 0, "positivacao": 0})
            categorias_dados[chave] = info

        # Industrializado/Thermo: mesma base de cálculo da planilha —
        # 25% do valor vendido é a meta industrializado, 3% é a meta
        # thermo; a comissão-teto é 1% e 2% em cima dessas metas.
        meta_industrializado = valor_vendido * 0.25
        meta_thermo = valor_vendido * 0.03
        industrializado_potencial = meta_industrializado * 0.01
        thermo_potencial = meta_thermo * 0.02

        rcas.append({
            "codigo": codigo,
            "nome": nome_rca.strip(),
            "rota": rota.strip(),
            "supervisor": supervisores.get(codigo, "SEM SUPERVISOR"),
            "valor_vendido": valor_vendido,
            "total_pedidos": total_pedidos,
            "categorias": categorias_dados,
            "industrializado_potencial": industrializado_potencial,
            "thermo_potencial": thermo_potencial,
            "premio_fixo": PREMIO_FIXO,
            "industrializado_participacao_pct": industrializado_pilares.get(codigo, {}).get("participacao_pct", 0),
            "industrializado_margem_pct": industrializado_pilares.get(codigo, {}).get("margem_pct", 0),
            "recompra_pct": recompra_por_rca.get(codigo, 0),
            "meta_posit_departamento": metas_departamento.get(codigo, {}),
        })

    metas_categoria_padrao = {chave: meta_posit for chave, _, _aba, _taxa, meta_posit in CATEGORIAS}
    taxas_categoria = {chave: taxa for chave, _, _aba, taxa, _meta_posit in CATEGORIAS}
    labels_categoria = {chave: label for chave, label, _aba, _taxa, _meta_posit in CATEGORIAS}

    constantes = {
        "dias_uteis": dias_uteis,
        "meta_pedidos_dia": meta_pedidos_dia,
        "taxa_padrao": taxa_padrao,
        "metas_categoria_padrao": metas_categoria_padrao,
        "taxas_categoria": taxas_categoria,
        "labels_categoria": labels_categoria,
        "ordem_categorias": [chave for chave, *_ in CATEGORIAS],
        "recompra_limite": RECOMPRA_LIMITE,
        "recompra_premio": RECOMPRA_PREMIO,
    }
    return rcas, constantes


def main():
    rcas, constantes = extrair()
    saida = {"rcas": rcas, "constantes": constantes}
    with open(CAMINHO_SAIDA, "w", encoding="utf-8") as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)
    print(f"{len(rcas)} RCAs extraídos. Salvo em: {CAMINHO_SAIDA}")


if __name__ == "__main__":
    main()
